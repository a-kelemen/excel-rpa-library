#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, generators, print_function, unicode_literals

from .exception import *
import os
import re
import string
from openpyxl.styles import colors
from openpyxl.utils.cell import get_column_letter
import xlwings as xw
from robot.libraries.BuiltIn import BuiltIn, RobotNotRunningError


class Workb(object):
	def __init__(self, book):
		self.wb_path = ""
		self.wb_name = ""
		self.wb_book = book
		self.folder = ""
		self.active_sheet = None

	def set_path(self, path):
		self.wb_path = path
		self.folder = os.path.dirname(path)

	def set_name(self, name):
		self.wb_name = name

	def set_active_sheet(self, sheet):
		self.active_sheet = sheet


class Base(object):
	_books = {}
	_default_folder = "C:/"

	def _get_path(self, path):
		if not os.path.exists(path):
			path = os.path.join(self._default_folder, path)
			if not os.path.exists(path):
				raise WorkbookNotFoundException(str("Workbook not found: ") + path)
		return path

	@staticmethod
	def _get_workbook(workbook):
		try:
			workbook = Base._books[workbook]
		except Exception:
			raise WorkbookNotFoundException("Workbook not found: " + workbook)
		return workbook

	@staticmethod
	def _get_coordinates_from_reference(ref):
		try:
			column_letters = re.search('^[A-Z]*', ref).group(0)
			row = re.search('[0-9]*$', ref).group(0)
		except AttributeError:
			raise Exception("Wrong reference!")
		num = 0
		for char in column_letters:
			if char in string.ascii_letters:
				num = 1 + num * 26 + (ord(char.upper()) - ord('A'))
		return int(num), int(row)

	def _select_sheet(self, workbook, selected_sheet):
		if selected_sheet == "active_":
			active_sheet = workbook.active_sheet
			if active_sheet is not None:
				return workbook.wb_book[active_sheet]
			else:
				raise WorksheetNotFoundException("No active sheet! Set active or add in argument!")
		else:
			for name in workbook.wb_book.sheetnames:
				if name == selected_sheet:
					return workbook.wb_book[name]
		raise WorksheetNotFoundException("Sheet not found: " + selected_sheet)

	def _is_formula(self, value):
		if str(value).startswith("="):
			return True
		else:
			return False

	def _get_formula_value(self, workbook, worksheet, row, col):
		app = xw.App(visible=False)
		wb = xw.Book(workbook)
		sht = None
		try:
			sht = wb.sheets[worksheet.title]
		except Exception:
			raise WorksheetNotFoundException("Sheet not found: " + worksheet)
		ref = self._to_reference(row, col)
		inter_value = sht.range(ref).value
		wb.close()
		app.kill()
		return inter_value

	def _nth_parent_folder(self, file_source, n):
		if n > 0:
			return self._nth_parent_folder(os.path.dirname(file_source), n-1)
		else:
			return file_source

	def _get_source(self, file_name):
		try:
			file_name = file_name.lstrip("\\").lstrip("/")
		except UnicodeDecodeError:
			pass
		if file_name == str("default_"):
			return self._get_process_folder()
		else:
			file_full_path = os.path.abspath(file_name)
			if not Base._is_absolute_path(file_name):
				return os.path.join(self._get_process_folder(), str(os.path.normpath(file_name)))
			else:
				return file_full_path

	@staticmethod
	def _is_absolute_path(path):
		return os.path.normpath(path) == os.path.abspath(path)

	def _get_process_folder(self):
		try:
			process_dir = os.path.dirname(BuiltIn().get_variable_value("${SUITE SOURCE}"))
			return process_dir
		except RobotNotRunningError as e:
			base_dir = self._nth_parent_folder(__file__, 3)
			return os.path.join(base_dir, str("tests"))

	@staticmethod
	def _get_color(color):
		if hasattr(colors, color.upper()):
			return getattr(colors, color.upper(), False)
		return None

	@staticmethod
	def _is_cell_empty(sheet, row, col):
		cell_value = sheet.cell(row=row, column=col).value
		if cell_value is None:
			return True
		return False

	@staticmethod
	def _to_reference(row, column):
		column_letter = get_column_letter(column)
		return column_letter + str(row)

	def _cell_right_to(self, reference):
		col, row = self._get_coordinates_from_reference(reference)
		col += 1
		return self._to_reference(row, col)

	def _cell_below(self, reference):
		col, row = self._get_coordinates_from_reference(reference)
		row += 1
		return self._to_reference(row, col)
