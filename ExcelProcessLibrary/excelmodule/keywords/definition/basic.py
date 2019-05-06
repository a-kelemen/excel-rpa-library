#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, generators, print_function, unicode_literals
from ..base import Base
from ..base import Workb
from ..exception import *

from openpyxl import load_workbook
import os


class BasicKeywords(Base):

	def open_workbook(self, name, alias=None):
		wb_source = self._get_source(name)
		if not os.path.exists(wb_source):
			raise WorkbookNotFoundException(str("Workbook not found: ") + wb_source)

		if alias is None and "default_" in Base._books:
			raise Exception("default workbook is already opened! set alias name to the workbook")

		if alias is not None:
			loaded = load_workbook(wb_source)
			wb = Workb(loaded)
			wb.set_path(wb_source)
			Base._books[alias] = wb
		elif alias is None or "default_" not in Base._books or Base._books["default_"] == "":
			loaded = load_workbook(wb_source)
			wb = Workb(loaded)
			wb.set_path(wb_source)
			Base._books["default_"] = wb

	def close_workbook(self, workbook):
		try:
			Base._books[workbook].wb_book.close()
			del Base._books[workbook]
		except KeyError:
			raise WorkbookNotFoundException(workbook + " workbook is not open.")

	def close_all_workbooks(self):
		for wbook in Base._books:
			Base._books[wbook].wb_book.close()
		Base._books = {}

	def create_worksheet(self, new_sheet_name, workbook):
		book = Base._get_workbook(workbook)
		exist = True
		try:
			self._select_sheet(book, new_sheet_name)
		except WorksheetNotFoundException:
			exist = False
		if not exist:
			book.wb_book.create_sheet(new_sheet_name)
			book.wb_book.save(book.wb_path)
		if exist:
			raise Exception(new_sheet_name + " worksheet already exist!")

	def delete_worksheet(self, sheet_name, workbook):
		book = Base._get_workbook(workbook)
		exist = False
		try:
			self._select_sheet(book, sheet_name)
			exist = True
		except WorksheetNotFoundException:
			pass
		if exist:
			if book.active_sheet == sheet_name:
				book.active_sheet = None
			book.wb_book.save(book.wb_path)
			std = book.wb_book[sheet_name]
			book.wb_book.remove(std)


		if not exist:
			raise WorksheetNotFoundException("Sheet not found: " + sheet_name)

	def set_active_worksheet(self, selected_sheet, workbook):
		book = Base._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)
		book.set_active_sheet(sheet.title)

	def rename_worksheet(self, selected_sheet, new_sheet_name, workbook):
		book = Base._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)
		exist = True
		try:
			self._select_sheet(book, new_sheet_name)
		except WorksheetNotFoundException:
			exist = False
		if not exist:
			sheet.title = new_sheet_name
			book.wb_book.save(book.wb_path)
		else:
			raise Exception(new_sheet_name + " worksheet already exist!")

	def save_workbook_as(self, new_workbook, workbook):
		book = Base._get_workbook(workbook)
		path_to_file = self._get_source(new_workbook)
		extension = os.path.splitext(path_to_file)[1]
		if extension == "":
			path_to_file += str(".xlsx")
		try:
			book.wb_book.save(path_to_file)
		except IOError:
			raise DirectoryNotFoundException(str("Directory not found: ") + path_to_file)
