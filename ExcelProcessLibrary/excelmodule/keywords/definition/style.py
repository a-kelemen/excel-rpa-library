#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, generators, print_function, unicode_literals
#from keywords.base import Base
from ..base import Base
from ..exception import *
from openpyxl import styles
from openpyxl.styles import Font, Color


class StyleKeywords(Base):

	def set_cell_background(self, reference, color, selected_sheet, workbook):
		col, row = Base._get_coordinates_from_reference(reference)
		book = self._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)

		fill_color = self._get_color(color)
		if fill_color is None:
			try:
				fill_color = Color(color)
			except ValueError as e:
				#Colors must be aRGB hex values
				print("ValueError:"+ e.__class__.__name__)#TODO
			#TODO raise Fail valueerror msg: Colors must be aRGB hex values or color names
		my_fill = styles.fills.PatternFill(patternType='solid', fgColor=fill_color)
		sheet.cell(row=row, column=col).fill = my_fill
		book.wb_book.save(book.wb_path)

	def set_font_color(self, reference, font_color, selected_sheet, workbook):
		col, row = Base._get_coordinates_from_reference(reference)
		book = self._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)

		fill_color = self._get_color(font_color)
		if fill_color is None:
			try:
				fill_color = Color(font_color)
			except ValueError as e:
				#Colors must be aRGB hex values
				print("ValueError:" + e.__class__.__name__ + "ez a color:" + font_color)
			#TODO raise Fail valueerror msg: Colors must be aRGB hex values or color names
		font = sheet.cell(row=row, column=col).font
		sheet.cell(row=row, column=col).font = Font(size=font.size,
													color=fill_color,
													bold=font.b,
													italic=font.i,
													underline=font.u)
		book.wb_book.save(book.wb_path)

	def set_cell_style(self, reference, style, selected_sheet, workbook):
		_styles = {"italic": True,
					"bold": True,
					"underline": "'single'"}
		col, row = Base._get_coordinates_from_reference(reference)
		book = self._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)
		if style.lower() in _styles:
			arg = style + "=" + str(_styles[style])
			font = sheet.cell(row=row, column=col).font
			if font.b and "bold" not in arg:
				arg += ", bold=True"
			if font.i and "italic" not in arg:
				arg += ", italic=True"
			if not font.u is None and "underline" not in arg:
				arg += ", underline='single'"
			arg += ", size=" + str(font.sz)
			if not font.color is None and not "rgb=None" in str(font.color):
				arg += ", color='" + str(font.color.rgb) + "'"
			sheet.cell(row=row, column=col).font = eval("Font(" + arg + ")")
		if style.lower() == "normal":
			sheet.cell(row=row, column=col).font = Font()
		book.wb_book.save(book.wb_path)

	def set_font_size(self, reference, font_size, selected_sheet, workbook):
		col, row = Base._get_coordinates_from_reference(reference)
		book = self._get_workbook(workbook)
		sheet = self._select_sheet(book, selected_sheet)
		font = sheet.cell(row=row, column=col).font
		sheet.cell(row=row, column=col).font = Font(size=font_size,
													color=font.color,
													bold=font.b,
													italic=font.i,
													underline=font.u)
		book.wb_book.save(book.wb_path)

	def remove_style(self, reference, selected_sheet, workbook):
		self.set_cell_style(reference, "normal", selected_sheet, workbook)

